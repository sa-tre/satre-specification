"""Create a SATRE evaluation CSV from the docs."""

from __future__ import annotations

import warnings
from os import path
from typing import TYPE_CHECKING, Any, Sequence

from docutils import nodes, writers
from docutils.frontend import OptionParser
from docutils.io import FileOutput

from sphinx import addnodes
from sphinx.builders import Builder
from sphinx.locale import __
from sphinx.util import logging
from sphinx.util.console import darkgreen  # type: ignore[attr-defined]
from sphinx.util.display import progress_message
from sphinx.util.nodes import inline_all_toctrees
from sphinx.util.osutil import ensuredir, make_filename_from_project
from sphinx.writers.text import TextTranslator, Table

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font

if TYPE_CHECKING:
    from sphinx.application import Sphinx
    from docutils.nodes import Element, Text


logger = logging.getLogger(__name__)


class SatreXlsxWriter(writers.Writer):
    """
    Write the extracted SATRE evaluation statements to an Excel file.
    """

    supported = ("text",)
    settings_spec = ("No options here.", "", ())
    settings_defaults: dict[str, Any] = {}

    output: bytes

    def __init__(self, builder: SatreCsvBuilder) -> None:
        super().__init__()
        self.builder = builder
        self.interested = None

    def translate(self) -> None:
        visitor = self.builder.create_translator(self.document, self.builder)
        self.document.walkabout(visitor)

        header_row = [
            "Section",
            "Item",
            "Statement",
            "Importance",
            "Score",
            "Response",
            "Improvements",
        ]
        column_widths = [10, 10, 10, 10, 10, 50, 50]
        rows = []
        for section in visitor.interested:
            title = section["section"][1][1].strip()
            # print(title)
            tables = section["table"]
            for table in tables:
                header = table.lines[0]
                # Is this a statement?
                if header[1].text.strip() == "Statement":
                    for line in table.lines[1:]:
                        number = line[0].text.strip()
                        statement = line[1].text.strip()
                        # guidance = line[2].strip()
                        importance = line[3].text.strip()
                        row = [title, number, statement, importance]
                        rows.append(row)
                        for i, cell in enumerate(row):
                            column_widths[i] = max(column_widths[i], len(str(cell)))

        wb = Workbook()
        ws = wb.active
        ws.append(header_row)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in rows:
            ws.append(row)

        for c, cell in enumerate(ws[1]):
            ws.column_dimensions[cell.column_letter].width = column_widths[c]

        buffer = BytesIO()
        wb.save(buffer)
        self.output = buffer.getvalue()


class SatreCsvTranslator(TextTranslator):
    """
    Extract SATRE evaluation statements from the docs and save as tabular file.

    sphinx.writers.text.TextTranslator converts a document into a single plain text file
    https://github.com/sphinx-doc/sphinx/blob/v7.2.5/sphinx/writers/text.py#L384-L1305

    This class extends TextTranslator to ignore everything apart from the relevant SATRE
    evaluation tables.
    """

    builder: SatreCsvBuilder

    def __init__(self, document: nodes.document, builder: SatreCsvBuilder) -> None:
        super().__init__(document, builder)
        self.interested = []

    # Disable default wrapping
    def end_state(
        self,
        wrap: bool = True,
        end: Sequence[str] | None = ("",),
        first: str | None = None,
    ) -> None:
        super().end_state(False, end, first)

    def visit_section(self, node: Element) -> None:
        self.new_state(0)
        self._title_char = self.sectionchars[self.sectionlevel]
        self.sectionlevel += 1
        if self.sectionlevel == 2:
            self.interested.append({"table": [], "section": None})

    def depart_section(self, node: Element) -> None:
        self.interested[-1]["section"] = self.states[-1][0]
        self.sectionlevel -= 1
        self.end_state()

    def visit_table(self, node: Element) -> None:
        if hasattr(self, "table"):
            msg = "Nested tables are not supported."
            # TODO: Outputting a warning here causes the CI check to fail since we use
            # SPHINXOPTS=-W to fail on errors. logging.skip_warningiserror doesn't work
            # :-(
            logger.info(msg)
            # with logging.skip_warningiserror(True):
            #     logger.warning(msg)
            # Ideally we'd raise an error since we shouldn't have any nested tables but
            # something is triggering this problem.
            # raise NotImplementedError(msg)
        self.new_state(0)
        self.table = Table()

    def depart_table(self, node: Element) -> None:
        self.interested[-1]["table"].append(self.table)
        self.add_text(str(self.table))
        # print(repr(self.table))
        del self.table
        self.end_state(wrap=False)

    def visit_start_of_file(self, node: Element) -> None:
        pass

    def depart_start_of_file(self, node: Element) -> None:
        pass


class SatreCsvBuilder(Builder):
    """
    Builds a SATRE evaluation TSV file.

    Based on https://github.com/sphinx-doc/sphinx/blob/v7.2.5/sphinx/builders/manpage.py#L29-L106
    """

    name = "satrecsv"
    format = "tsv"
    epilog = __("The output is in %(outdir)s.")

    default_translator_class = SatreCsvTranslator
    supported_image_types: list[str] = []

    def init(self) -> None:
        # section numbers for headings in the currently visited document
        self.secnumbers: dict[str, tuple[int, ...]] = {}

    def get_outdated_docs(self) -> str | list[str]:
        return "all documents"

    def get_target_uri(self, docname: str, typ: str | None = None) -> str:
        return ""

    @progress_message(__("writing"))
    def write(self, *ignored: Any) -> None:
        docwriter = SatreXlsxWriter(self)
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=DeprecationWarning)
            # DeprecationWarning: The frontend.OptionParser class will be replaced
            # by a subclass of argparse.ArgumentParser in Docutils 0.21 or later.
            docsettings: Any = OptionParser(
                defaults=self.env.settings,
                components=(docwriter,),
                read_config_files=True,
            ).get_default_values()

        docname = self.config.root_doc

        # targetname = make_filename_from_project(self.config.project) + ".xlsx"
        targetname = "satre.xlsx"

        logger.info(darkgreen(targetname) + " { ", nonl=True)
        destination = FileOutput(
            destination_path=path.join(self.outdir, targetname), encoding="utf-8"
        )

        tree = self.env.get_doctree(docname)
        docnames: set[str] = set()
        largetree = inline_all_toctrees(
            self, docnames, docname, tree, darkgreen, [docname]
        )
        largetree.settings = docsettings
        logger.info("} ", nonl=True)
        self.env.resolve_references(largetree, docname, self)
        # remove pending_xref nodes
        for pendingnode in largetree.findall(addnodes.pending_xref):
            pendingnode.replace_self(pendingnode.children)

        docwriter.write(largetree, destination)


def setup(app: Sphinx) -> dict[str, Any]:
    """
    Register the extension as a builder.
    """
    app.add_builder(SatreCsvBuilder)

    return {
        "version": "0.0.1",
        "parallel_read_safe": True,
        "parallel_write_safe": True,
    }
