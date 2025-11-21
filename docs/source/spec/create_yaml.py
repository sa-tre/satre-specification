import math
import openpyxl
import pandas as pd
from ruamel.yaml import YAML

# Read the Excel file
df = pd.read_excel("SATRE.2.0.-.Reorg.xlsx")

# Load workbook to extract hyperlinks
wb = openpyxl.load_workbook("SATRE.2.0.-.Reorg.xlsx")
sheets = wb.sheetnames
ws = wb[sheets[0]]
assert ws.cell(1,4).value == "Link to Architecture View"

urls = []
for n in range(1, 161):
    c = ws.cell(n + 1, 4)
    if c.value and c.hyperlink:
        print(c.hyperlink.target)
        urls.append(c.hyperlink.target)
    else:
        urls.append(None)

# Convert rows to dictionaries
def row_to_dict(r):
    d = {
        "pillar": r["Pillar"],
        "capability_index": str(r["Capability No."]),
        "requirement_index": r["Requirement No"].rstrip("."),
        "statement": r["Statement"],
        "guidance": r["Guidance"],
        "importance": r["Importance"],
    }
    return d

is_na = df.isnull().apply(lambda x: all(x), axis=1) 
statements = df[~is_na]
statements = statements[~statements["Pillar"].str.match("Version")]
requirements = statements.apply(row_to_dict, axis=1)
requirements = list(requirements)

# Add architecture URLs
assert len(requirements) == len(urls)
for n in range(len(urls)):
    requirements[n]["architecture_url"] = urls[n]

# Write to YAML
yaml = YAML()
yaml.default_flow_style = False
yaml.indent(mapping=2, sequence=4, offset=2)
with open("specification.yaml", "w", encoding="utf-8") as f:
    yaml.dump({"specification": requirements}, f)

print(f"Created specification.yaml with {len(requirements)} requirements")
