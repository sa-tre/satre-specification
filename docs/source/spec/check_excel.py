import openpyxl
import pandas as pd

# Load the Excel file
df = pd.read_excel("SATRE.2.0.-.Reorg.xlsx")
print(f"Total rows in dataframe: {len(df)}")
print(f"Columns: {list(df.columns)}")
print("\nFirst few rows:")
print(df.head())

# Check hyperlinks
wb = openpyxl.load_workbook("SATRE.2.0.-.Reorg.xlsx")
sheets = wb.sheetnames
print(f"\nSheet names: {sheets}")

ws = wb[sheets[0]]
print(f"\nHeader row (row 1):")
for col in range(1, 8):
    print(f"  Col {col}: {ws.cell(1, col).value}")

print(f"\nChecking hyperlinks in column 4 (Link to Architecture View):")
print(f"Cell D1 value: {ws.cell(1, 4).value}")

# Check first 10 data rows for hyperlinks
for row in range(2, 12):
    cell = ws.cell(row, 4)
    req_cell = ws.cell(row, 3)  # Requirement No column
    has_hyperlink = cell.hyperlink is not None
    print(f"Row {row} (Req: {req_cell.value}): Value='{cell.value}', Has hyperlink: {has_hyperlink}")
    if has_hyperlink:
        print(f"  -> URL: {cell.hyperlink.target}")
