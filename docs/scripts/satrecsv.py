import csv
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import os
from subprocess import check_output, CalledProcessError
import argparse

MAX_COLUMN_WIDTH = 70


def get_version() -> str:
    """
    Obtain a version to use in documentation.

    If this is readthedocs use the RTD environment variables and the git SHA,
    otherwise lookup the git version
    """
    try:
        git_sha = check_output(["git", "rev-parse", "HEAD"]).strip().decode()[:7]
    except (CalledProcessError, FileNotFoundError):
        git_sha = "unknown"

    # https://docs.readthedocs.io/en/stable/reference/environment-variables.html#envvar-READTHEDOCS_VERSION
    rtd_version_slug = os.getenv("READTHEDOCS_VERSION")
    rtd_version_type = os.getenv("READTHEDOCS_VERSION_TYPE")

    if rtd_version_slug and rtd_version_type:
        if rtd_version_type == "tag":
            return f"{rtd_version_slug}-{git_sha}"
        else:
            return f"{rtd_version_type}-{rtd_version_slug}-{git_sha}"
    else:
        try:
            return check_output(["git", "describe", "--always"]).strip().decode()
        except (CalledProcessError, FileNotFoundError):
            return "unknown"


def main():
    parser = argparse.ArgumentParser(
        description="Create a SATRE evaluation CSV/XLSX from the YAML spec."
    )
    parser.add_argument(
        "--output-dir", default="build/satrecsv", help="Directory to save the outputs"
    )
    parser.add_argument(
        "--input-file", default="source/spec/specification.yaml", help="Input YAML file"
    )
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)

    with open(args.input_file, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    header_row = [
        "Section",
        "Item",
        "Statement",
        "Guidance",
        "Importance",
        "Score",
        "Response",
        "Improvements",
    ]

    rows = []

    for item in data.get("specification", []):
        row = [
            item["pillar"],
            item["requirement_index"],
            item["statement"],
            (item.get("guidance") or "").strip(),
            item["importance"],
            "",  # Score
            "",  # Response
            "",  # Improvements
        ]
        rows.append(row)

    version = get_version()

    # Write CSV
    csv_path = os.path.join(args.output_dir, "satre.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header_row)
        for row in rows:
            writer.writerow(row)
        writer.writerow([])
        writer.writerow(["Version", version])

    # Write XLSX
    xlsx_path = os.path.join(args.output_dir, "satre.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.append(header_row)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    column_widths = [10, 10, 10, 10, 10, 10, 50, 50]
    for row in rows:
        ws.append(row)
        for i, cell in enumerate(row):
            column_widths[i] = min(
                max(column_widths[i], len(str(cell))), MAX_COLUMN_WIDTH
            )

    for c, cell in enumerate(ws[1]):
        ws.column_dimensions[cell.column_letter].width = column_widths[c]

    ws.append([])
    ws.append(["Version", version])

    # Set word-wrap on all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(xlsx_path)
    print(f"Written {csv_path} and {xlsx_path}")


if __name__ == "__main__":
    main()
